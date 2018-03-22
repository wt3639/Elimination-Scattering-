import openpyxl
import sys
import random
from openpyxl import load_workbook
from openpyxl import Workbook

sum_book = Workbook()
sum_sheet = sum_book.active
sum_sheet["B1"].value = 'Ⅰ'
sum_sheet["C1"].value = 'Ⅱ'
sum_sheet["D1"].value = 'Ⅲ'
sum_sheet["E1"].value = 'Ⅳ'
sum_sheet["F1"].value = 'Ⅴ'
for t in range(1,len(sys.argv)):
    #output = open(r'C:\Users\wutao\Desktop\point', 'w+')
    print(sys.argv[t]+' start processing...')
    xlsfile = sys.argv[t] # 打开指定路径中的xls文件
    book = load_workbook(xlsfile)#得到Excel文件的book对象，实例化对象
    sheet0 = book.active # 通过sheet索引获得sheet对象
    nrows = sheet0.max_row    # 获取行总数
    print("row",nrows)
    ncols = sheet0.max_column   #获取列总数
    print("col",ncols)
    row_1 = sheet0[1]    # 获得第1行的数据列表
    col_1 = sheet0["A"]    # 获得第1列的数据列表
    leftest = 7 
    for i in range(1,nrows):
        for j in range(1,ncols):
            #if(float(sheet0.cell_value(i,j)) < 0):
                 #change_sheet.write(i,j,0)
            if(j>leftest and j<31 and i==1):
                for l in range(0,(j-leftest)):
                    max_num = sheet0.cell(row=j-leftest+1+1,column = j+1).value
                    min_num = 0
                    writeValue = min_num+(max_num-min_num)/(j-leftest+1)*(l+1)
                    sheet0.cell(row=i+l+1,column=j+1).value = writeValue
                    #output.write(str(i+l)+' '+str(j)+'\n')
            if(row_1[j].value==col_1[i].value and j>30):
                for k in range(0,20):
                    min_num = sheet0.cell(row=i-4+1,column = j+1).value
                    max_num = sheet0.cell(row=i+17+1,column = j+1).value
                    writeValue = min_num+(max_num-min_num)/21*(k+1)
                    sheet0.cell(row=i-3+k+1,column=j+1).value = writeValue
                    #output.write(str(i-3+k)+' '+str(j)+'\n')
            if(row_1[j].value*2 == col_1[i].value and i>1):
                for n in range(0,24):
                    max_num = sheet0.cell(row=i-6+1,column=j+1).value
                    if(i+18+1>nrows):
                        min_num = 0
                    else:
                        min_num = sheet0.cell(row=i+18+1,column=j+1).value
                    if(i-5+n<nrows):
                        writeValue = max_num-(max_num-min_num)/25*(n+1)
                        sheet0.cell(row=i-5+n+1,column=j+1).value = writeValue
                        #output.write(str(i-5+n)+' '+str(j)+'\n')
                if(i>=nrows-2):
                    for m in range(j+1,ncols):
                        if(i-5+m-j<=nrows):
                            for n in range(0,24):
                                max_num = sheet0.cell(row=i-6+m-j+1,column=m+1).value
                                if(i+18+1+m-j>nrows):
                                    min_num = 0
                                else:
                                    min_num = sheet0.cell(row=i+18+m-j+1,column=m+1).value 
                                if(i-5+n+m-j<nrows):
                                    writeValue = max_num-(max_num-min_num)/25*(n+1)
                                    sheet0.cell(row=i-5+n+m-j+1,column=m+1).value = writeValue
                                    #output.write(str(i-5+n+m-j)+' '+str(m)+'\n')
    for row in sheet0.iter_rows(min_col=2,min_row=2, max_col=ncols, max_row=nrows):
        for cell in row:
            #print(cell.value)
            if(cell.value <0):
                 cell.value = 0    
                 
    for i in range(1,nrows):
        for j in range(1,ncols):
            if(sheet0.cell(row=i+1,column=j+1).value == 0):
                rown = i
                coln = j
                if(rown-1==0):
                    up_rown = rown
                else:
                    up_rown = rown - 1
                if(rown+1 == nrows):
                    down_rown = rown
                else:
                    down_rown = rown + 1
                if(coln-1 == 0):
                    left_coln = coln
                else:
                    left_coln = coln - 1
                if(coln+1 == ncols):
                    right_coln = coln
                else:
                    right_coln = coln + 1
    
                right_up =sheet0.cell(row=up_rown+1,column=right_coln+1).value
                right =sheet0.cell(row=rown+1,column=right_coln+1).value
                right_down=sheet0.cell(row=down_rown+1,column=right_coln+1).value
                up=sheet0.cell(row=up_rown+1,column=coln+1).value
                down =sheet0.cell(row=down_rown+1,column=coln+1).value
                left_up=sheet0.cell(row=up_rown+1,column=left_coln+1).value
                left=sheet0.cell(row=rown+1,column=left_coln+1).value
                left_down=sheet0.cell(row=down_rown+1,column=left_coln+1).value
                middle = (right_up+right+right_down+up+down+left_up+left+left_down)/8
                #print(i,j,right_up,right,right_down,up,down,left,left_up,left_down,middle)
                sheet0.cell(row=i+1,column=j+1).value = middle
    sum1=0
    sum2=0
    sum3=0
    sum4=0
    sum5=0
    #part1 sum
    for row in sheet0.iter_rows(min_col=2,min_row=2, max_col=27, max_row=42):
        for cell in row:
            sum1 = sum1 + cell.value
    #part2 sum
    for row in sheet0.iter_rows(min_col=2,min_row=43, max_col=27, max_row=67):
        for cell in row:
            sum2 = sum2 + cell.value
    #part3 sum
    for row in sheet0.iter_rows(min_col=2,min_row=68, max_col=27, max_row=nrows):
        for cell in row:
            sum3 = sum3 + cell.value
    #part4 sum
    for row in sheet0.iter_rows(min_col=28,min_row=2, max_col=ncols, max_row=67):
        for cell in row:
            sum4 = sum4 + cell.value
    #part5 sum
    for row in sheet0.iter_rows(min_col=28,min_row=68, max_col=ncols, max_row=nrows):
        for cell in row:
            sum5 = sum5 + cell.value
    sum_sheet["A%d"%(t+1)].value = sys.argv[t].split('\\')[len(sys.argv[t].split('\\'))-1]
    sum_sheet["B%d"%(t+1)].value = sum1
    sum_sheet["C%d"%(t+1)].value = sum2
    sum_sheet["D%d"%(t+1)].value = sum3
    sum_sheet["E%d"%(t+1)].value = sum4
    sum_sheet["F%d"%(t+1)].value = sum5
    save_file = sys.argv[t].split('.xlsx')[0]+'(Elimination Scattering).xlsx'
    book.save(save_file)
    print(sys.argv[t]+' processed completely')
#output.close()
sum_file=''
for i in range(len(sys.argv[1].split('\\'))-1):
    sum_file = sum_file+sys.argv[1].split('\\')[i]+'\\'
print(sum_file)
sum_book.save(sum_file+'Fluorescence Regional Integration.xlsx')
print('All files complete!')



